"""
Excel export/import celej agendy.

Export:
- Automaticky exportuje VŠETKY používateľské tabuľky z databázy
- Každá tabuľka do vlastného sheetu
- Dropdown listy pre číselníkové hodnoty (typ_dokladu, forma_uhrady, mena, sadzba_dph)
- Možnosť vlastnej hodnoty v dropdownoch

Import:
- Načítanie dát zo všetkých sheetov
- Automaticky nájde zodpovedajúcu tabuľku v databáze
- Validácia a import do databázy
"""

import sqlite3
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


# === ŠTÝLY ===
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

# === DROPDOWN HODNOTY ===
DROPDOWN_VALUES = {
    'typ_dokladu_prijem': ['Faktúra vystavená', 'Pokladničný doklad', 'Bankový výpis', 'Interný doklad', 'Iné'],
    'typ_dokladu_vydavok': ['Faktúra prijatá', 'Pokladničný doklad výdavok', 'Bankový výpis výdavok', 'Interný doklad', 'Iné'],
    'forma_uhrady': ['Hotovosť', 'Bankový prevod', 'Karta', 'Šek', 'Iné'],
    'mena': ['EUR', 'USD', 'CZK', 'GBP', 'CHF'],
    'sadzba_dph': ['0', '5', '19', '20', '23'],
    'stav': ['neuhradena', 'ciastocne_uhradena', 'uhradena', 'po_splatnosti'],
    'typ_odvodu': ['Sociálne poistenie', 'Zdravotné poistenie', 'Dôchodkové poistenie', 'Iné'],
    'boolean': ['0', '1'],
    'druh_majetku': ['HM', 'NM'],
}

# === MAPOVANIE SHEET → TABUĽKA ===
SHEET_TO_TABLE = {
    'Nastavenia': 'nastavenia',
    'Prijmy': 'prijmy',
    'Prijmy_polozky': 'prijmy_polozky',
    'Vydavky': 'vydavky',
    'Vydavky_polozky': 'vydavky_polozky',
    'Majetok': 'majetok',
    'Zasoby': 'zasoby',
    'Pohladavky': 'pohladavky',
    'Zavazky': 'zavazky',
    'Odvody': 'odvody',
    'Adresar': 'adresar',
    'Adresar_dorucovacie_adresy': 'adresar_dorucovacie_adresy',
    'Adresar_kontakty': 'adresar_kontakty',
    'Adresar_bankove_ucty': 'adresar_bankove_ucty',
    'Adresar_poznamky': 'adresar_poznamky',
    'Ciselniky': 'ciselniky_dokladov',
    'Sablony': 'sablony_poloziek',
    'Jednotky': 'jednotky',
    'Objednavky': 'objednavky',
    'Objednavky_polozky': 'objednavky_polozky',
}

# === DROPDOWN MAPOVANIE PODĽA TABUĽKY ===
TABLE_DROPDOWNS = {
    'prijmy': {
        'typ_dokladu': 'typ_dokladu_prijem',
        'forma_uhrady': 'forma_uhrady',
        'mena': 'mena',
        'sadzba_dph': 'sadzba_dph',
        'danovy_prijem': 'boolean',
    },
    'vydavky': {
        'typ_dokladu': 'typ_dokladu_vydavok',
        'forma_uhrady': 'forma_uhrady',
        'mena': 'mena',
        'sadzba_dph': 'sadzba_dph',
        'danovy_vydavok': 'boolean',
    },
    'prijmy_polozky': {
        'sadzba_dph': 'sadzba_dph',
    },
    'vydavky_polozky': {
        'sadzba_dph': 'sadzba_dph',
    },
    'majetok': {
        'druh_majetku': 'druh_majetku',
    },
    'pohladavky': {
        'mena': 'mena',
        'stav': 'stav',
    },
    'zavazky': {
        'mena': 'mena',
        'stav': 'stav',
    },
    'odvody': {
        'typ_odvodu': 'typ_odvodu',
    },
    'adresar': {
        'platca_dph': 'boolean',
        'je_aktivny': 'boolean',
    },
    'adresar_dorucovacie_adresy': {
        'je_aktivny': 'boolean',
    },
    'adresar_kontakty': {
        'je_aktivny': 'boolean',
    },
    'adresar_bankove_ucty': {
        'je_aktivny': 'boolean',
    },
    'adresar_poznamky': {
        'je_aktivny': 'boolean',
    },
    'ciselniky_dokladov': {
        'typ_dokladu': 'typ_dokladu_prijem',
        'je_aktivny': 'boolean',
    },
    'sablony_poloziek': {
        'typ_dokladu': 'typ_dokladu_prijem',
        'sadzba_dph': 'sadzba_dph',
        'je_aktivny': 'boolean',
    },
    'jednotky': {
        'je_aktivny': 'boolean',
    },
    'objednavky': {
        'stav': 'stav',
        'mena': 'mena',
    },
    'objednavky_polozky': {
        'sadzba_dph': 'sadzba_dph',
    },
}

# === PORADIE EXPORTU/IMPORTU (závislosti) ===
# Tabuľky bez cudzích kľúčov najprv, závislé neskôr
IMPORT_ORDER = [
    'nastavenia', 'jednotky', 'zobrazenie',
    'adresar', 'ciselniky_dokladov', 'sablony_poloziek',
    'prijmy', 'vydavky', 'majetok', 'zasoby',
    'pohladavky', 'zavazky', 'odvody', 'objednavky',
    'prijmy_polozky', 'vydavky_polozky', 'objednavky_polozky',
    'adresar_dorucovacie_adresy', 'adresar_kontakty',
    'adresar_bankove_ucty', 'adresar_poznamky',
]


def _style_header(cell):
    """Nastaví štýl hlavičky."""
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGNMENT
    cell.border = THIN_BORDER


def _style_data_cell(cell):
    """Nastaví štýl dátovej bunky."""
    cell.border = THIN_BORDER


def _add_dropdown(ws, column_letter, values, start_row=2, allow_blank=True):
    """Pridá data validation dropdown do stĺpca."""
    dv = DataValidation(
        type="list",
        formula1='"' + ','.join(values) + '"',
        allow_blank=allow_blank,
        showDropDown=False
    )
    dv.error = 'Vyberte hodnotu zo zoznamu alebo zadajte vlastnú'
    dv.errorTitle = 'Neplatná hodnota'
    dv.prompt = 'Vyberte zo zoznamu alebo zadajte vlastnú hodnotu'
    dv.promptTitle = 'Výber hodnoty'
    ws.add_data_validation(dv)
    # Aplikuj na celý stĺpec (max 10000 riadkov)
    dv.add(f'{column_letter}{start_row}:{column_letter}10000')


def _get_table_data(cursor, table_name):
    """Získa všetky dáta z tabuľky ako list dictov."""
    try:
        cursor.execute(f"SELECT * FROM {table_name}")
        rows = cursor.fetchall()
        columns = [description[0] for description in cursor.description]
        return [dict(zip(columns, row)) for row in rows]
    except sqlite3.OperationalError:
        return []


def _write_sheet(wb, sheet_name, data, columns=None, dropdowns=None, summary_cols=None):
    """Vytvorí sheet a zapíše doň dáta s hlavičkami."""
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(title=sheet_name)

    # Urči hlavičky — z dát alebo zo zadaných stĺpcov
    if data:
        headers = list(data[0].keys())
    elif columns:
        headers = columns
    else:
        ws.append(['Žiadne dáta'])
        return ws

    # Hlavičky
    ws.append(headers)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        _style_header(cell)

    # Dáta
    for row_data in data:
        row_values = []
        for header in headers:
            val = row_data.get(header)
            if val is None:
                row_values.append('')
            else:
                row_values.append(val)
        ws.append(row_values)
        row_idx = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            _style_data_cell(ws.cell(row=row_idx, column=col_idx))

    # Sumárne riadky (pre číselné stĺpce)
    if summary_cols and data:
        _add_summary_rows(ws, headers, data, summary_cols)

    # Auto-fit šírky stĺpcov
    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)

    # Dropdowny
    if dropdowns:
        for col_name, dropdown_key in dropdowns.items():
            if col_name in headers:
                col_idx = headers.index(col_name) + 1
                col_letter = get_column_letter(col_idx)
                values = DROPDOWN_VALUES.get(dropdown_key, [])
                if values:
                    _add_dropdown(ws, col_letter, values)

    # Zamrznutá hlavička
    ws.freeze_panes = 'A2'

    return ws


def _add_summary_rows(ws, headers, data, summary_cols):
    """Pridá sumárne riadky na koniec sheetu."""
    from openpyxl.styles import Font, PatternFill, Alignment

    # Prázdny riadok
    ws.append([''] * len(headers))

    # Nadpis sumára
    sum_row = ws.max_row + 1
    ws.cell(row=sum_row, column=1, value='SUMÁR')
    ws.cell(row=sum_row, column=1).font = Font(bold=True, size=12, color='FFFFFF')
    ws.cell(row=sum_row, column=1).fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    ws.cell(row=sum_row, column=1).alignment = Alignment(horizontal='left')

    # Spočítaj sumy pre každý summary stĺpec
    for col_name in summary_cols:
        if col_name in headers:
            col_idx = headers.index(col_name) + 1
            total = 0
            for row_data in data:
                val = row_data.get(col_name)
                if val is not None:
                    try:
                        total += float(val)
                    except (ValueError, TypeError):
                        pass
            sum_row = ws.max_row + 1
            ws.cell(row=sum_row, column=1, value=f'Spočet {col_name}:')
            ws.cell(row=sum_row, column=col_idx, value=round(total, 2))
            ws.cell(row=sum_row, column=1).font = Font(bold=True)
            ws.cell(row=sum_row, column=col_idx).font = Font(bold=True)
            ws.cell(row=sum_row, column=col_idx).number_format = '#,##0.00'

    # Počet záznamov
    sum_row = ws.max_row + 1
    ws.cell(row=sum_row, column=1, value=f'Počet záznamov: {len(data)}')
    ws.cell(row=sum_row, column=1).font = Font(bold=True)


def _create_summary_sheet(wb, cursor):
    """Vytvorí sumárny sheet s prehľadom príjmov a výdavkov."""
    from openpyxl.styles import Font, PatternFill, Alignment

    ws = wb.create_sheet(title='Sumarizacia', index=1)

    # === PRÍJMY PODĽA MESIACOV ===
    ws.append(['PRÍJMY PODĽA MESIACOV'])
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    ws.merge_cells('A1:D1')

    ws.append(['Mesiac', 'Počet', 'Celková suma', 'Celková DPH'])
    for cell in ws[3]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    cursor.execute("""
        SELECT 
            strftime('%Y-%m', datum_prijetia) as mesiac,
            COUNT(*) as pocet,
            COALESCE(SUM(suma), 0) as celkova_suma,
            COALESCE(SUM(dph), 0) as celkova_dph
        FROM prijmy
        GROUP BY mesiac
        ORDER BY mesiac
    """)
    for row in cursor.fetchall():
        ws.append([row['mesiac'], row['pocet'], row['celkova_suma'], row['celkova_dph']])

    # === VÝDAVKY PODĽA MESIACOV ===
    ws.append([''])
    ws.append(['VÝDAVKY PODĽA MESIACOV'])
    start_row = ws.max_row
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=14, color='FFFFFF')
    ws.cell(row=start_row, column=1).fill = PatternFill(start_color='C62828', end_color='C62828', fill_type='solid')
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)

    ws.append(['Mesiac', 'Počet', 'Celková suma', 'Celková DPH'])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    cursor.execute("""
        SELECT 
            strftime('%Y-%m', datum_uhrady) as mesiac,
            COUNT(*) as pocet,
            COALESCE(SUM(suma), 0) as celkova_suma,
            COALESCE(SUM(dph), 0) as celkova_dph
        FROM vydavky
        GROUP BY mesiac
        ORDER BY mesiac
    """)
    for row in cursor.fetchall():
        ws.append([row['mesiac'], row['pocet'], row['celkova_suma'], row['celkova_dph']])

    # === PRÍJMY PODĽA TYPU DOKLADU ===
    ws.append([''])
    ws.append(['PRÍJMY PODĽA TYPU DOKLADU'])
    start_row = ws.max_row
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=14, color='FFFFFF')
    ws.cell(row=start_row, column=1).fill = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)

    ws.append(['Typ dokladu', 'Počet', 'Celková suma', 'Celková DPH'])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    cursor.execute("""
        SELECT 
            typ_dokladu,
            COUNT(*) as pocet,
            COALESCE(SUM(suma), 0) as celkova_suma,
            COALESCE(SUM(dph), 0) as celkova_dph
        FROM prijmy
        GROUP BY typ_dokladu
        ORDER BY celkova_suma DESC
    """)
    for row in cursor.fetchall():
        ws.append([row['typ_dokladu'], row['pocet'], row['celkova_suma'], row['celkova_dph']])

    # === VÝDAVKY PODĽA TYPU DOKLADU ===
    ws.append([''])
    ws.append(['VÝDAVKY PODĽA TYPU DOKLADU'])
    start_row = ws.max_row
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=14, color='FFFFFF')
    ws.cell(row=start_row, column=1).fill = PatternFill(start_color='C62828', end_color='C62828', fill_type='solid')
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)

    ws.append(['Typ dokladu', 'Počet', 'Celková suma', 'Celková DPH'])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    cursor.execute("""
        SELECT 
            typ_dokladu,
            COUNT(*) as pocet,
            COALESCE(SUM(suma), 0) as celkova_suma,
            COALESCE(SUM(dph), 0) as celkova_dph
        FROM vydavky
        GROUP BY typ_dokladu
        ORDER BY celkova_suma DESC
    """)
    for row in cursor.fetchall():
        ws.append([row['typ_dokladu'], row['pocet'], row['celkova_suma'], row['celkova_dph']])

    # === POHĽADÁVKY A ZÁVÄZKY ===
    ws.append([''])
    ws.append(['POHĽADÁVKY A ZÁVÄZKY'])
    start_row = ws.max_row
    ws.cell(row=start_row, column=1).font = Font(bold=True, size=14, color='FFFFFF')
    ws.cell(row=start_row, column=1).fill = PatternFill(start_color='1565C0', end_color='1565C0', fill_type='solid')
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=4)

    ws.append(['Kategória', 'Počet', 'Celková suma', 'Celková DPH'])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    cursor.execute("""
        SELECT 
            'Pohľadávky' as kategoria,
            COUNT(*) as pocet,
            COALESCE(SUM(suma), 0) as celkova_suma,
            COALESCE(SUM(dph), 0) as celkova_dph
        FROM pohladavky
        UNION ALL
        SELECT 
            'Záväzky' as kategoria,
            COUNT(*) as pocet,
            COALESCE(SUM(suma), 0) as celkova_suma,
            COALESCE(SUM(dph), 0) as celkova_dph
        FROM zavazky
    """)
    for row in cursor.fetchall():
        ws.append([row['kategoria'], row['pocet'], row['celkova_suma'], row['celkova_dph']])

    # Formátovanie číselných stĺpcov
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value != 0:
                cell.number_format = '#,##0.00'

    # Auto-fit šírky
    for col_idx in range(1, 5):
        max_length = 15
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 40)

    return ws


def _get_user_tables(cursor):
    """Získa zoznam všetkých používateľských tabuliek z databázy."""
    cursor.execute("""
        SELECT name FROM sqlite_master
        WHERE type='table'
          AND name NOT LIKE 'sqlite_%'
          AND name NOT LIKE 'schema_%'
          AND name != 'migrations'
        ORDER BY name
    """)
    return [row[0] for row in cursor.fetchall()]


def _table_to_sheet_name(table_name):
    """Konvertuje názov tabuľky na názov sheetu (Title Case, bez podtržnítk)."""
    # Použi známe mapovanie ak existuje
    for sheet, table in SHEET_TO_TABLE.items():
        if table == table_name:
            return sheet
    # Fallback: prvé písmeno veľké, podtržnítko → žiadna medzera
    return table_name.replace('_', ' ').title().replace(' ', '')


def _sheet_to_table_name(sheet_name):
    """Konvertuje názov sheetu na názov tabuľky."""
    return SHEET_TO_TABLE.get(sheet_name, sheet_name.lower().replace(' ', '_'))


def export_agenda_to_excel(db_path, output_path=None, nazov_agendy=None):
    """
    Exportuje celú agendu do Excel súboru.
    Automaticky exportuje VŠETKY používateľské tabuľky z databázy.

    Returns:
        output_path: cesta k vytvorenému súboru
    """
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    wb = Workbook()
    # Odstrániť default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Získať všetky používateľské tabuľky
    tables = _get_user_tables(cursor)

    for table_name in tables:
        data = _get_table_data(cursor, table_name)
        sheet_name = _table_to_sheet_name(table_name)
        dropdowns = TABLE_DROPDOWNS.get(table_name)
        # Ak nie sú dáta, získaj stĺpce z PRAGMA pre hlavičky
        columns = None
        if not data:
            try:
                cursor.execute(f"PRAGMA table_info({table_name})")
                columns = [row[1] for row in cursor.fetchall()]
            except sqlite3.OperationalError:
                columns = None
        # Sumárne stĺpce pre príjmy a výdavky
        summary_cols = None
        if table_name == 'prijmy':
            summary_cols = ['suma', 'dph']
        elif table_name == 'vydavky':
            summary_cols = ['suma', 'dph']
        elif table_name == 'pohladavky':
            summary_cols = ['suma', 'dph']
        elif table_name == 'zavazky':
            summary_cols = ['suma', 'dph']
        elif table_name == 'objednavky':
            summary_cols = ['suma', 'dph', 'celkova_suma']
        _write_sheet(wb, sheet_name, data, columns=columns, dropdowns=dropdowns, summary_cols=summary_cols)

    # === SUMARIZAČNÝ SHEET ===
    _create_summary_sheet(wb, cursor)

    # === INFO SHEET ===
    info_ws = wb.create_sheet(title='Info', index=0)
    info_ws.append(['Export agendy - Daňová evidencia'])
    info_ws.append(['Dátum exportu:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    info_ws.append([''])
    info_ws.append(['Pokyny:'])
    info_ws.append(['- Neupravujte názvy sheetov ani hlavičky stĺpcov'])
    info_ws.append(['- Pre import upravte dáta a uložte súbor'])
    info_ws.append(['- Stĺpce s dropdownom umožňujú vlastnú hodnotu'])
    info_ws.append(['- ID stĺpce sa pri importe ignorujú (autoincrement)'])
    info_ws.append(['- Pre vymazanie riadku ho označte žltou farbou v stĺpci A'])
    info_ws.append([''])
    info_ws.append(['Tabuľky:'])
    for sheet_name in wb.sheetnames:
        if sheet_name not in ('Info', 'Sumarizacia'):
            info_ws.append([f'- {sheet_name}'])
    info_ws.append([''])
    info_ws.append(['Sumarizacia:'])
    info_ws.append(['- Prehľad príjmov a výdavkov podľa mesiacov'])
    info_ws.append(['- Prehľad podľa typov dokladov'])
    info_ws.append(['- Prehľad pohľadávok a záväzkov'])

    info_ws['A1'].font = Font(bold=True, size=14)

    conn.close()

    if output_path is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        # Použi explicitný názov agendy, inak názov súboru DB
        if nazov_agendy is None:
            nazov_agendy = os.path.splitext(os.path.basename(db_path))[0]
        # Sanitizácia názvu pre súbor
        nazov_agendy = nazov_agendy.replace(' ', '_').replace('.', '_')
        output_path = os.path.join(
            os.path.dirname(db_path),
            f'{nazov_agendy}_export_{timestamp}.xlsx'
        )

    wb.save(output_path)
    return output_path


def _read_sheet_data(ws):
    """Načíta dáta zo sheetu ako list dictov."""
    if ws.max_row < 1:
        return []

    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value))
        else:
            headers.append('')

    if not headers or headers[0] == 'Žiadne dáta':
        return []

    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or v == '' for v in row):
            continue
        row_dict = {}
        for idx, header in enumerate(headers):
            if idx < len(row):
                val = row[idx]
                # Konvertuj None na prázdny string pre textové polia
                if val is None:
                    val = ''
                row_dict[header] = val
            else:
                row_dict[header] = ''
        data.append(row_dict)

    return data


def import_agenda_from_excel(excel_path, db_path, mode='merge'):
    """
    Importuje dáta z Excelu do databázy.
    Automaticky nájde zodpovedajúce tabuľky pre každý sheet.

    Args:
        excel_path: cesta k Excel súboru
        db_path: cesta k databáze
        mode: 'merge' (pridá k existujúcim) alebo 'replace' (vymaže existujúce)

    Returns:
        (success, message, stats)
    """
    wb = load_workbook(excel_path, data_only=True)

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    stats = {}

    # Získať zoznam existujúcich tabuliek v databáze
    cursor.execute("""
        SELECT name FROM sqlite_master
        WHERE type='table'
          AND name NOT LIKE 'sqlite_%'
          AND name NOT LIKE 'schema_%'
          AND name != 'migrations'
    """)
    existing_tables = {row[0] for row in cursor.fetchall()}

    # Vypnúť foreign key constraints počas importu
    cursor.execute("PRAGMA foreign_keys = OFF")

    # Zoradiť sheety podľa import poradia (závislosti)
    sheet_order = []
    for table in IMPORT_ORDER:
        for sheet, tbl in SHEET_TO_TABLE.items():
            if tbl == table and sheet in wb.sheetnames:
                sheet_order.append(sheet)
                break
    # Pridať zvyšné sheety (neznáme poradie)
    for sheet in wb.sheetnames:
        if sheet != 'Info' and sheet not in sheet_order:
            sheet_order.append(sheet)

    if mode == 'replace':
        # Vymaž dáta v opačnom poradí (najprv závislé)
        for sheet_name in reversed(sheet_order):
            table = _sheet_to_table_name(sheet_name)
            if table not in existing_tables:
                continue
            try:
                cursor.execute(f"DELETE FROM {table}")
                stats[f"{table}_deleted"] = cursor.rowcount
            except sqlite3.OperationalError:
                pass

    total_inserted = 0

    for sheet_name in sheet_order:
        if sheet_name == 'Info':
            continue

        table = _sheet_to_table_name(sheet_name)
        if table not in existing_tables:
            stats[f"{sheet_name}_skipped"] = 'Tabuľka neexistuje v databáze'
            continue

        ws = wb[sheet_name]
        data = _read_sheet_data(ws)

        if not data:
            continue

        try:
            # Získaj existujúce stĺpce
            cursor.execute(f"PRAGMA table_info({table})")
            existing_columns = {row[1] for row in cursor.fetchall()}

            # Filtrované dáta - iba stĺpce ktoré existujú, bez 'id'
            filtered_rows = []
            for row in data:
                filtered = {}
                for k, v in row.items():
                    if k in existing_columns and k != 'id':
                        # Konvertuj prázdne stringy na None pre číselné polia
                        if v == '' and k in ['suma', 'dph', 'zaklad_dane', 'celkova_suma',
                                              'mnozstvo', 'jednotkova_cena_bez_dph',
                                              'zostatkova_cena', 'rocny_odpis', 'odpisova_skupina',
                                              'uhradena_suma', 'sadzba_dph_zakladna',
                                              'sadzba_dph_znizena', 'sadzba_dph_super_znizena',
                                              'predcislie', 'cislo_uctu']:
                            filtered[k] = None
                        else:
                            filtered[k] = v
                if filtered:
                    filtered_rows.append(filtered)

            if not filtered_rows:
                continue

            # Vlož dáta
            columns = list(filtered_rows[0].keys())
            placeholders = ', '.join(['?' for _ in columns])
            columns_str = ', '.join(columns)

            inserted = 0
            for row in filtered_rows:
                values = []
                for col in columns:
                    val = row.get(col)
                    if val == '' and col in ['created_at', 'updated_at']:
                        values.append(datetime.now().isoformat())
                    else:
                        values.append(val)
                try:
                    cursor.execute(
                        f"INSERT INTO {table} ({columns_str}) VALUES ({placeholders})",
                        values
                    )
                    inserted += 1
                except sqlite3.IntegrityError:
                    # Duplicita - preskoč
                    pass
                except sqlite3.OperationalError as e:
                    stats[f"{table}_error"] = str(e)
                    break

            stats[f"{table}_inserted"] = inserted
            total_inserted += inserted

        except sqlite3.OperationalError as e:
            stats[f"{table}_error"] = str(e)

    # Zapnúť foreign key constraints späť
    cursor.execute("PRAGMA foreign_keys = ON")

    conn.commit()
    conn.close()
    wb.close()

    message = f"Importovaných {total_inserted} záznamov"
    return True, message, stats
